import os

import openpyxl
import requests


def channel_excel():
    file_path = os.path.join("Excel", "channel.xlsx")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Channels_list']

    data = []

    row = 3
    while any(sheet.cell(row=row, column=col).value for col in range(1, 10)):
        if sheet.cell(row=row, column=2).value == 'TCP':
            entry = {
                "name": f"{sheet.cell(row=row, column=1).value}" or "",
                "type": f"{sheet.cell(row=row, column=2).value}" or "",
                "ip": f"{sheet.cell(row=row, column=3).value}" or "",
                "port": f"{sheet.cell(row=row, column=4).value}" or "",
                "active": f"{sheet.cell(row=row, column=5).value}" or 0,
            }
            url = "http://127.0.0.1:7777/channel/"
            response = requests.post(url, json=entry)
            if response.status_code == 200:
                print("Запрос успешно выполнен!")
                print("Ответ сервера:", response.json())
            else:
                print("Ошибка при выполнении запроса. Код:", response.status_code)
            row += 1
        elif sheet.cell(row=row, column=2).value == 'COM':
            entry = {
                "name": f"{sheet.cell(row=row, column=1).value}" or "",
                "type": f"{sheet.cell(row=row, column=2).value}" or "",
                "serial_port": f"{sheet.cell(row=row, column=6).value}" or "",
                "baud_rate": f"{sheet.cell(row=row, column=7).value}" or "",
                "data_bits": f"{sheet.cell(row=row, column=8).value}" or "",
                "stop_bits": f"{sheet.cell(row=row, column=9).value}" or "",
                "active": f"{sheet.cell(row=row, column=5).value}" or 0,
            }
            url = "http://127.0.0.1:7777/channel/"
            response = requests.post(url, json=entry)
            if response.status_code == 200:
                print("Запрос успешно выполнен!")
                print("Ответ сервера:", response.json())
            else:
                print("Ошибка при выполнении запроса. Код:", response.status_code)
            row += 1

    wb.close()
