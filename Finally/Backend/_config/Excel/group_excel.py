import os

import openpyxl
import requests


def group_excel():
    file_path = os.path.join("Excel", "group.xlsx")

    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Group_list']

    data = []

    row = 3
    while any(sheet.cell(row=row, column=col).value for col in range(1, 10)):
        entry = {
            "group_name": f"{sheet.cell(row=row, column=1).value}" or "",
        }
        url = "http://127.0.0.1:7777/group/"
        response = requests.post(url, json=entry)
        if response.status_code == 200:
            print("Запрос успешно выполнен!")
            print("Ответ сервера:", response.json())
        else:
            print("Ошибка при выполнении запроса. Код:", response.status_code)
        row += 1

    wb.close()
