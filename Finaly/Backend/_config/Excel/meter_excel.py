import os

import openpyxl
import requests


def meter_excel():
    file_path = os.path.join("Excel", "meter.xlsx")

    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Meter_list']

    data = []

    row = 3
    while any(sheet.cell(row=row, column=col).value for col in range(1, 10)):

        entry = {
            "meter_name": f"{sheet.cell(row=row, column=1).value}" or "",
            "meter_serial": f"{sheet.cell(row=row, column=3).value}" or "",
            "channel_name": f"{sheet.cell(row=row, column=5).value}" or "",
            "group_name": [group.strip() for group in f"{sheet.cell(row=row, column=9).value}".split(',')] or [],
            "active": f"{sheet.cell(row=row, column=8).value}" or "",
            "physical_address": f"{sheet.cell(row=row, column=4).value}" or "",
            "protocol_name": f"{sheet.cell(row=row, column=2).value}" or "",
            "ki": f"{sheet.cell(row=row, column=6).value}" or "",
            "ku": f"{sheet.cell(row=row, column=7).value}" or "",
        }
        print(entry)
        url = "http://127.0.0.1:7777/meter/"
        response = requests.post(url, json=entry)
        if response.status_code == 200:
            print("Запрос успешно выполнен!")
            print("Ответ сервера:", response.json())
        else:
            print("Ошибка при выполнении запроса. Код:", response.status_code)
        row += 1

    wb.close()
